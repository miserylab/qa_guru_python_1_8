import csv
from zipfile import ZipFile
import os
import os.path
import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook


@pytest.fixture(scope='session')
def dir_check():
    if not os.path.exists('resources'):
        os.makedirs('resources')


@pytest.fixture(scope='session')
def file_remove():
    if not os.path.exists('tmp'):
        os.makedirs('tmp')
    yield
    os.remove('./tmp/docs-pytest-org-en-latest.pdf')
    os.remove('./tmp/file_example_XLSX_50.xlsx')
    os.remove('./tmp/username.csv')
    os.rmdir('tmp')


def get_file_extension(filename):
    root, extension = os.path.splitext(filename)
    return extension


def test_zip_files(dir_check, file_remove):
    zip_ = ZipFile('resources/sample.zip', 'w')

    zip_.write('docs-pytest-org-en-latest.pdf')
    zip_.write('file_example_XLSX_50.xlsx')
    zip_.write('username.csv')
    fileNames = zip_.namelist()
    for filename in fileNames:
        if get_file_extension(filename) == '.pdf':
            zip_.extract(filename, 'tmp')
            reader = PdfReader(filename)
            number_of_pages = len(reader.pages)
            page = reader.pages[0]
            text = page.extract_text()
            assert "pytest Documentation" in text
            assert number_of_pages == 412
            assert os.path.getsize(filename) == 1739253
        elif get_file_extension(filename) == '.xlsx':
            zip_.extract(filename, 'tmp')
            workbook = load_workbook(filename)
            sheet = workbook.active
            assert 'Mara' in sheet.cell(row=3, column=2).value
            assert os.path.getsize(filename) == 7360
        elif get_file_extension(filename) == '.csv':
            zip_.extract(filename, 'tmp')
            with open(filename, encoding="utf8") as f:
                reader = csv.DictReader(f, delimiter=";")
                list_f = list(reader)
                assert "Rachel" in (list_f[0]['First name'])
                assert os.path.getsize(filename) == 183
        else:
            print(filename, "n/a")
    zip_.close()
